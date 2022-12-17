#coding:utf-8
#since 2022-02-12

import streamlit as st
import numpy as np
import psycopg2
import openpyxl
import base64
from pikepdf import Pdf
from pyecharts import options as opts
from pyecharts.charts import Map
import streamlit.components.v1 as components

def chemicalcompliance():
    def findillegalchar(casnum):
        '''
        Detect illegal char in the inputted casnum to prevent attack 
        Return None if not.
        '''
        safetynum=['0','1','2','3','4','5','6','7','8','9','-']
        for char in casnum:
            if char not in safetynum:            
                return char

    class Query:
        '''
        :use single cas num to search regulation
        :return a class 'numpy.ndarray'
        '''
        def __init__(self,cas_number):
            self.__casno=cas_number

        def CasnumberQuery(self):
            #There are 5 tables (IECSC,TSCAINV,PMNACC,EC and CNOTHERS) have casno field.
            #query in CNOTHERS
            result_list_0=[]
            row=0
            CAS_X=str(self.__casno)
            cas_query_0='''SELECT ori_sn,casno,cnname,enname,remark,legid FROM CNOTHERS WHERE casno=%s;'''
            #1336-21-6为氨的水溶液（7664-41-7为氨气、液氨）
            chemicals.execute(cas_query_0,(CAS_X,))
            result_0=chemicals.fetchall()
            for chem in result_0:#unpackaged to list
                result_list_0.append(chem[5])
            result_set_0=set(result_list_0)#convert list to set
            if len(result_set_0)==0:#if nothong in the set,it means no any matched result
                return None
            else:
                leg_query='''SELECT leg_cn,leg_en,pub_date FROM CNLAWS WHERE legid='LEG_X';'''
                leg_arr=np.empty((len(result_set_0)+1,3),object)#leg_cn|leg_en|pub_date
                leg_arr[0,:]=(['法规中文名称','English Title','发布日期'])
                for leg_id in result_set_0:
                    leg_query_1=leg_query.replace('LEG_X',leg_id)
                    chemicals.execute(leg_query_1)
                    leg_result_1=chemicals.fetchone()
                    row+=1
                    leg_arr[row,:]=([leg_result_1[0],leg_result_1[1],leg_result_1[2]])
                return leg_arr

    class BatchQuery:
        '''
        :received an uploaded xlsx file
        :use a list of cas num in xlsx file to search regulations
        :return a xlsx file(to do)
        :return an array to screen(for now)
        :return illegal char if any
        :as numpy array can't be expended or concatenated in place(make sense because arr always occupied a contiguous block of memory). rewrite the code from array to list
        '''
        def __init__(self,xlsx_file):
            self.__xlsx_file=xlsx_file
        
        def b_query(self):
            workbook_0=openpyxl.load_workbook(self.__xlsx_file)#目标文件
            sheet_0=workbook_0.active
            #sheet_0.insert_cols(2,3)#第2列前插入3列（法规中文名称|法规英文名称|发布日期）
            rownum=0
            rowmax=sheet_0.max_row
            chem_i=''
            x=[]
            illgealchar=''        
            arr_titlerow=[['CAS No.','CN Title','EN Title','Publish Date']]
            result_truck_2=[]
            for rownum in range(1,rowmax+1,1):
                result_truck=[]
                result_truck_1=[]
                result_truck_2=[]
                chem_i=sheet_0.cell(row=rownum,column=1).value
                
                if rowmax==1 and chem_i is None:#???为什么xlsx文件时空白的但是max_row检测出来的结果仍然是1
                    st.warning('您上传的文件是空白的，请检查！')
                    return None
                elif findillegalchar(chem_i):
                    arr_titlerow.append([chem_i,'非法字符：'+illgealchar,'illegal character detected!','illegal character detected!'])
                else: 
                    query_truck=Query(chem_i)
                    result_truck=query_truck.CasnumberQuery()
                    if result_truck is None:
                        arr_titlerow.append([chem_i,'No match','No match','No match'])
                    else:
                        rownum_truck=result_truck.shape[0]#取得数组的行数值            
                        arr_newcol=np.full(rownum_truck,chem_i,order='F')#构建数组只有1列且每行赋值为chem_i（cas no）
                        result_truck_1=np.column_stack((arr_newcol,result_truck))#将单列数组与2维数组合并（arr_newcol位置在前）
                        result_truck_2=result_truck_1.tolist()#ndarray转换为list
                        for x in result_truck_2[1:]:#解包
                            arr_titlerow.append(x)
                        del result_truck
                        del result_truck_1
                        del result_truck_2
            arr_batch=np.array(arr_titlerow)
            return arr_batch

    
    st.title('化学品关联法规查询')
    st.title('Find Chemical in Which Regulation')
    #warehouse=psycopg2.connect(**st.secrets["postgres"])
    #chemicals=warehouse.cursor()
    st.markdown('**Bad_news!** Free Heroku Postgres service would be ended since 2022-11-28. Pls check [offical announcement of Heroku](https://blog.heroku.com/next-chapter).')
    st.info('单个查询 Single Query')
    whichcasno=st.text_input('输入CAS号码 Enter CAS number', value='', max_chars=None, key=None, type='default', help='CAS num looks like 1336-21-6')#how to use the key？
    st.text('or')
    st.info('批量查询 Bulk Query')
    uploaded_xlsx_file=st.file_uploader('上传文件格式为xlsx且第一列为CAS no',type=['xlsx'])

    if whichcasno!='':
        if findillegalchar(whichcasno):
            st.write(whichcasno,'包含非法字符：',findillegalchar(whichcasno))
        else:
            st.write(whichcasno,'的关联法规为：')
            query_test=Query(whichcasno)
            df_result_0=query_test.CasnumberQuery()
            if df_result_0 is not None:
                st.dataframe(data=df_result_0) #此处应可调单元格宽度和高度
            else:
                st.warning('未检索到关联法规数据 no any match!')   

    if uploaded_xlsx_file is not None:#如果有文件上传
        bquery_test=BatchQuery(uploaded_xlsx_file) 
        bq_result_1=bquery_test.b_query()
        if bq_result_1 is not None:

            #***********************************************************************************
            xlsx_exporter=openpyxl.Workbook()
            sheet=xlsx_exporter.active

            i=0
            rownum_final=bq_result_1.shape[0]
            for i in range(0,rownum_final,1):
                    sheet.cell(row=i+1,column=1).value=bq_result_1[i,0]#'CAS No.'
                    sheet.cell(row=i+1,column=2).value=bq_result_1[i,1]#'CN Title'
                    sheet.cell(row=i+1,column=3).value=bq_result_1[i,2]#'EN Title'
                    sheet.cell(row=i+1,column=4).value=bq_result_1[i,3]#'Publish Date'              
            xlsx_exporter.save('results.xlsx')
            data=open('results.xlsx','rb').read()#readonly in binary format
            b64 = base64.b64encode(data).decode('UTF-8')
            href = f'<a href="data:file/data;base64,{b64}" download="results.xlsx">Download xlsx file</a>'
            st.markdown(href, unsafe_allow_html=True)
            xlsx_exporter.close()
            st.dataframe(data=bq_result_1)
            #***********************************************************************************
    return None

def unlock_file():
    '''
    Function: remove content restrictions set by user password but not owner password from your pdf file 
    '''
    st.title('清除上传PDF文件的内容限制')
    st.header('Remove content restrictions from the uploaded PDF file')
    uploaded_pdf_file=st.file_uploader('Upload your PDF file',type=['pdf'])#It said you will get a 'StringIO object' now

    if uploaded_pdf_file is not None:
        pdffile = Pdf.open(uploaded_pdf_file)
        pdffile.save('unlocked.pdf') 
        with open("unlocked.pdf", "rb") as pdf_file:#Thanks Franky1 from Streamlit community
            PDFbyte = pdf_file.read()
            st.download_button(label="Download unlocked.pdf",
                data=PDFbyte,
                file_name="unlocked.pdf",
                mime='application/octet-stream')

def geo_map():
    st.header('可视化数据地图（省份）生成工具')
    st.subheader('Visualization Data Map(Province) Online Creator')
    st.markdown('[下载XLSX模板文件 Download XLSX Template File](https://mypsdata.s3.us-east-2.amazonaws.com/template.xlsx)')#链接来自onedrive
    uploaded_geodata_file=st.file_uploader('注：通过上传XLSX文件来设置地图的地名变量、标题',type=['xlsx'])
    if uploaded_geodata_file is not None:
        workbook_geodata=openpyxl.load_workbook(uploaded_geodata_file,read_only=True)
        sheet_geodata=workbook_geodata.active
        geodata_list=[]
        mymaintitle=''
        mysubtitle='Total：'
        myseriestitle=''
        docsum=0#各省发布文件数量总计
        for rownum in range(1,35,1):#read data of each province
            geodata_i=()#建立空元组
            geodata_i=(sheet_geodata.cell(row=rownum,column=1).value,sheet_geodata.cell(row=rownum,column=2).value)
            docsum=docsum+sheet_geodata.cell(row=rownum,column=2).value
            geodata_list.append(geodata_i)
        mymaintitle=sheet_geodata.cell(row=35,column=2).value
        mysubtitle=f'{mysubtitle}{docsum}'
        myseriestitle=sheet_geodata.cell(row=36,column=2).value
        docnumber_china_map = (Map(init_opts=opts.InitOpts(
            width='1024px',
            height='768px',
            page_title = "Data Map"))
            .add(series_name=myseriestitle,
            data_pair=geodata_list,
            maptype="china",
            is_roam=True,
            is_map_symbol_show=False)
            .set_global_opts(
                title_opts=opts.TitleOpts(title=mymaintitle,subtitle=mysubtitle,pos_right='center'),
                visualmap_opts=opts.VisualMapOpts(max_=100, is_piecewise=True),      
                legend_opts=opts.LegendOpts(is_show=False)) 
            .set_series_opts(label_opts=opts.LabelOpts(is_show=False))
            .render_embed()
            ) 
        workbook_geodata.close()
        #docnumber_china_map.render('中国新法规地图.html')
        components.html(docnumber_china_map,width=1024, height=768)
    else:
        #*********************map(default)********************
        uploaded_geodata_file= [('西藏',30),('新疆',39),('河南',45),('北京',22),('河北',10),('辽宁',12),('江西',6),('上海',69),('安徽',59),('江苏',16),('湖南',9),('浙江',36),('海南',2),('广东',22),('湖北',8),('黑龙江',11),('澳门',1),('陕西',11),('四川',7),('内蒙古',3),('重庆',3),('云南',6),('贵州',2),('吉林',3),('山西',12),('山东',11),('福建',4),('青海',1),('香港',57)]
        docnum_source_date=f'2022年xx月x日-xx日期间共有xxxx个法规文件发布'
        docnumber_china_map = (Map(init_opts=opts.InitOpts(
            width='1024px',
            height='768px',
            page_title = "新法规地图"))
            .add(series_name='新法规文件发布数量',
            data_pair=uploaded_geodata_file,
            maptype="china",
            is_roam=True,
            is_map_symbol_show=False)
            .set_global_opts(
                title_opts=opts.TitleOpts(title='中国各省新法规发布数量（Sample）',subtitle=docnum_source_date,pos_right='center'),
                visualmap_opts=opts.VisualMapOpts(max_=100, is_piecewise=True),      
                legend_opts=opts.LegendOpts(is_show=False)) 
            .set_series_opts(label_opts=opts.LabelOpts(is_show=False))
            .render_embed()
            ) 
        components.html(docnumber_china_map,width=1024, height=768)
        #**********************************************************************    
          
            
def sidebar():
    whichtool=st.sidebar.radio("Choose Compliance Tool:",('Chemical → Regulation','Online Map Creator','Remove PDF Password'))

    return whichtool

def main():
    st.set_page_config(page_title="Compliance Bridge",layout="wide")#2021-05-16
    toolnum=sidebar()
    if toolnum=='Online Map Creator':
        geo_map()   
    if toolnum=='Chemical → Regulation':#Note:There are space between strings!
        chemicalcompliance()
    if toolnum=='Remove PDF Password':
        unlock_file() 
    with st.sidebar:
        st.markdown('  ')
        st.markdown('  ')
        st.markdown('  ')
        st.markdown('  ')
        st.markdown('  ')
        st.markdown('  ')
        st.markdown('  ')
        st.markdown('  ')
        st.markdown('  ')
        st.markdown('  ')
        st.markdown('  ')
        st.markdown('  ')
        st.markdown('  ')
        st.markdown('  ')
        st.markdown('  ')
        st.markdown('  ')
        st.markdown('  ')
        st.markdown('  ')
        st.markdown('  ')
        st.markdown('  ')
        st.markdown('  ')
        st.markdown('  ')
        st.markdown('  ')
        st.markdown('  ')
        st.markdown('  ')
        st.markdown('  ')
        st.markdown('  ')
        st.markdown('  ')
        st.markdown('''	:copyright:Some copyrights may be reserved so''')#20220417add copyrignt symbol
        st.markdown('''Please contact  [Yenan Chen](mailto:bchen@nimonik.com)''')
    return None

if __name__ == '__main__':
    main()
